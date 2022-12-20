from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import openpyxl
# from openpyxl import workbook
from datetime import datetime
import os

FilePath = "C:\\Users\\MSI\\Documents\\GitHub\\Selenium-Python\\Data\\data_test.xlsx"
dataSheet = "Sheet1"
wb = openpyxl.load_workbook(FilePath)
ws = wb[dataSheet]

today = datetime.now()
parent_dir = "C:\\Users\\MSI\\Documents\\GitHub\\Selenium-Python\\Data\\Result"
#------Create diractory--------------
path = os.path.join(parent_dir, "Result_" + today.strftime('%Y-%m-%d_%H.%M.%S'))
os.mkdir(path)

row_count = ws.max_row
col_count = ws.max_column
print("row is : ", row_count, "column is : ", col_count)

driver = webdriver.Chrome(executable_path="c:\\browserdivers\\chromedrivers.exe")
driver.get("https://pypi.org/")
driver.maximize_window()
time.sleep(2)
print(driver.title)

for x in range(row_count-1):
    x = x + 2
    login = driver.find_element(By.XPATH, '//*[@id="user-indicator"]/nav[1]/ul/li[3]/a')
    login.click()
    time.sleep(2)
    print(x)
    username = driver.find_element(By.XPATH, '//*[@id="username"]')
    username.send_keys(ws.cell(x, 3).value)
    time.sleep(2)

    password = driver.find_element(By.XPATH, '//*[@id="password"]')
    password.send_keys(ws.cell(x, 4).value)
    time.sleep(2)

    btn_login = driver.find_element(By.XPATH, '//*[@id="content"]/div/div/form/div[2]/div[3]/div/div/input')
    btn_login.click()
    time.sleep(2)

    if ws.cell(x, 1).value == "TC_001":
        validate_login = driver.find_element(By.XPATH, '//*[@id="username-errors"]/ul/li')
    elif ws.cell(x, 1).value == "TC_002":
        validate_login = driver.find_element(By.XPATH, '//*[@id="password-errors"]/ul/li')
        
    Data_check = validate_login.text
    time.sleep(2)
    
    #------check passed and failed--------
    if Data_check == ws.cell(x, 6).value:
        ws.cell(x, 5).value = "Passed"
        # ws(FilePath, "_", ds, "Sheet1", x, 5, )
    else:   
        ws.cell(x, 5).value = "Failed"
    wb.save (path + "\\Result_" + today.strftime('%Y-%m-%d_%H.%M.%S') + ".xlsx")

driver.close()